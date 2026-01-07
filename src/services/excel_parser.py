"""Excel order file parser for multi-worksheet order files."""

from __future__ import annotations
from io import BytesIO
from typing import Optional
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import BaseModel, Field


class ExcelOrderItem(BaseModel):
    """A single item from an Excel order."""
    category: str = Field(description="Category from worksheet name")
    subcategory: Optional[str] = Field(default=None, description="Subcategory column")
    product_name: str = Field(description="Product name")
    unit: str = Field(description="Unit of measurement")
    price: Optional[float] = Field(default=None, description="Unit price")
    quantity: float = Field(description="Order quantity (Opening Order)")
    row_number: int = Field(description="Row number in the worksheet")


class ExcelOrderSheet(BaseModel):
    """Parsed data from a single worksheet (category)."""
    category: str
    items: list[ExcelOrderItem] = Field(default_factory=list)
    total_items: int = 0
    total_value: Optional[float] = None


class ExcelOrderResult(BaseModel):
    """Complete result of parsing an Excel order file."""
    success: bool
    filename: Optional[str] = None
    customer_name: Optional[str] = None
    sheets: list[ExcelOrderSheet] = Field(default_factory=list)
    total_items: int = 0
    total_categories: int = 0
    total_value: Optional[float] = None
    warnings: list[str] = Field(default_factory=list)
    error: Optional[str] = None


# Common column name variations
COLUMN_MAPPINGS = {
    "subcategory": ["subcategory", "sub-category", "sub category", "type", "subcat"],
    "product_name": ["product", "product name", "item", "item name", "description", "name"],
    "unit": ["unit", "uom", "unit of measure", "measure", "units"],
    "price": ["price", "unit price", "rate", "cost", "amount"],
    "quantity": ["opening order", "order", "qty", "quantity", "order qty", "order quantity", "amount"],
}


def find_column_index(headers: list[str], column_type: str) -> Optional[int]:
    """Find the column index for a given column type based on common name variations."""
    variations = COLUMN_MAPPINGS.get(column_type, [])
    for i, header in enumerate(headers):
        if header and header.lower().strip() in variations:
            return i
    return None


def parse_worksheet(sheet: Worksheet, category: str) -> ExcelOrderSheet:
    """Parse a single worksheet into order items."""
    items = []
    warnings = []

    # Get headers from first row
    headers = []
    for cell in sheet[1]:
        headers.append(str(cell.value).lower().strip() if cell.value else "")

    # Find column indices
    subcategory_idx = find_column_index(headers, "subcategory")
    product_idx = find_column_index(headers, "product_name")
    unit_idx = find_column_index(headers, "unit")
    price_idx = find_column_index(headers, "price")
    quantity_idx = find_column_index(headers, "quantity")

    # Product name is required
    if product_idx is None:
        # Try to find any column that might be product name
        for i, h in enumerate(headers):
            if h and "product" in h.lower():
                product_idx = i
                break

    if product_idx is None:
        return ExcelOrderSheet(
            category=category,
            items=[],
            total_items=0,
        )

    # Parse data rows (skip header)
    total_value = 0.0
    for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        # Skip empty rows
        if not row or all(cell is None or cell == "" for cell in row):
            continue

        # Get product name
        product_name = row[product_idx] if product_idx < len(row) else None
        if not product_name:
            continue

        # Get quantity - this is crucial
        quantity = None
        if quantity_idx is not None and quantity_idx < len(row):
            qty_value = row[quantity_idx]
            if qty_value is not None:
                try:
                    quantity = float(qty_value)
                except (ValueError, TypeError):
                    pass

        # Skip rows without quantity (no order)
        if quantity is None or quantity <= 0:
            continue

        # Get other fields
        subcategory = None
        if subcategory_idx is not None and subcategory_idx < len(row):
            subcategory = row[subcategory_idx]
            if subcategory:
                subcategory = str(subcategory).strip()

        unit = "units"
        if unit_idx is not None and unit_idx < len(row):
            unit_value = row[unit_idx]
            if unit_value:
                unit = str(unit_value).strip()

        price = None
        if price_idx is not None and price_idx < len(row):
            price_value = row[price_idx]
            if price_value is not None:
                try:
                    price = float(price_value)
                    total_value += price * quantity
                except (ValueError, TypeError):
                    pass

        items.append(ExcelOrderItem(
            category=category,
            subcategory=subcategory,
            product_name=str(product_name).strip(),
            unit=unit,
            price=price,
            quantity=quantity,
            row_number=row_num,
        ))

    return ExcelOrderSheet(
        category=category,
        items=items,
        total_items=len(items),
        total_value=total_value if total_value > 0 else None,
    )


def parse_excel_order(
    file_content: bytes,
    filename: Optional[str] = None,
    customer_name: Optional[str] = None,
) -> ExcelOrderResult:
    """
    Parse a multi-worksheet Excel order file.

    Args:
        file_content: Raw bytes of the Excel file
        filename: Original filename (optional)
        customer_name: Customer name if known (optional)

    Returns:
        ExcelOrderResult with parsed order data
    """
    try:
        # Load workbook from bytes
        workbook = load_workbook(filename=BytesIO(file_content), read_only=True, data_only=True)
    except Exception as e:
        return ExcelOrderResult(
            success=False,
            filename=filename,
            error=f"Failed to read Excel file: {str(e)}",
        )

    sheets = []
    total_items = 0
    total_value = 0.0
    warnings = []

    # Process each worksheet as a category
    for sheet_name in workbook.sheetnames:
        # Skip hidden sheets or metadata sheets
        if sheet_name.lower() in ["metadata", "config", "settings", "info"]:
            continue

        sheet = workbook[sheet_name]

        # Use sheet name as category
        category = sheet_name.strip()

        parsed_sheet = parse_worksheet(sheet, category)

        if parsed_sheet.items:
            sheets.append(parsed_sheet)
            total_items += parsed_sheet.total_items
            if parsed_sheet.total_value:
                total_value += parsed_sheet.total_value

    workbook.close()

    if not sheets:
        return ExcelOrderResult(
            success=False,
            filename=filename,
            error="No valid order items found in the Excel file. Make sure worksheets have 'Product' and 'Opening Order' columns.",
        )

    return ExcelOrderResult(
        success=True,
        filename=filename,
        customer_name=customer_name,
        sheets=sheets,
        total_items=total_items,
        total_categories=len(sheets),
        total_value=total_value if total_value > 0 else None,
        warnings=warnings,
    )


def excel_order_to_text(result: ExcelOrderResult) -> str:
    """
    Convert parsed Excel order to a text format for LLM processing.

    This allows the existing extraction pipeline to process Excel orders
    using the same confirmation and routing logic.
    """
    if not result.success:
        return ""

    lines = []

    if result.customer_name:
        lines.append(f"Order from: {result.customer_name}")
    if result.filename:
        lines.append(f"File: {result.filename}")
    lines.append("")

    for sheet in result.sheets:
        lines.append(f"=== {sheet.category} ===")
        for item in sheet.items:
            line = f"- {item.product_name}: {item.quantity} {item.unit}"
            if item.subcategory:
                line = f"- [{item.subcategory}] {item.product_name}: {item.quantity} {item.unit}"
            if item.price:
                line += f" @ {item.price}"
            lines.append(line)
        lines.append("")

    if result.total_value:
        lines.append(f"Total estimated value: {result.total_value:,.2f}")

    return "\n".join(lines)
